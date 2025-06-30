from time import sleep
import random
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font
import re


def clean_digit(x):
    if not x:
        return None
    cleaned = x.replace('$', '').replace(',', '')
    return float(cleaned)

with (sync_playwright() as p):
    browser = p.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://www.amazon.com", timeout=60000)
    b = page.query_selector('button[alt="Continue shopping"]')
    if b: b.click()
    page.wait_for_selector("input#twotabsearchtextbox")
    page.fill("input#twotabsearchtextbox", "wireless headphones")
    page.press("input#twotabsearchtextbox", "Enter")
    page.wait_for_load_state("load")


    wb = Workbook()
    ws = wb.active
    ws.title = "amazon listing of products"

    headers = ["Name", "Price", "Number of sellers"]
    ws.append(headers)
    ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = ws.column_dimensions['C'].width = 20

    pp = context.new_page()
    while True:
      for el in page.query_selector_all('div[role="listitem"]'):
        if not el.query_selector("a"):
          break
        pp.goto("https://www.amazon.com/" + el.query_selector("a").get_attribute("href"))
        pp.wait_for_load_state("load")
        sleep(random.uniform(1, 2))
        name = pp.query_selector('span[id="productTitle"]').inner_text()
        price = pp.query_selector(".a-price.aok-align-center > .a-offscreen")
        pricestr = price.inner_text().replace('$', '') if price else ""
        price = float(pricestr) if pricestr else 0.0
        nos = 1
        a = pp.query_selector('a[id="aod-ingress-link"]')
        if a:
          a.click()
          h5 = 'h5[id="aod-filter-offer-count-string"]'
          try:
            pp.wait_for_selector(h5)
            nos += int(re.search(r'\d+', pp.query_selector(h5).text_content()).group())
          except:
              pass
        ws.append([name, price, nos])

      next_link = page.locator('span.s-pagination-strip > ul > li').last.locator("span > a")
      if next_link.is_visible():
        sleep(random.uniform(4, 8))
        next_link.click()
        sleep(random.uniform(4, 8))
      else:
        break
wb.save("sample_data.xlsx")